#!/usr/bin/env python3
"""
Emergency Backup Script
========================
Moves sensitive data files out of the repo tree into ~/Documents/bwsync/.

Creates an Excel workbook with:
  - Tab 1: Chrome profile inventory (from CSV)
  - Tab 2: Metadata (date, source counts)

If BWSYNC_EXCEL_PASSWORD env var is set, encrypts the workbook with
msoffcrypto-tool (Office-standard encryption).

Also copies passwords.zip as-is (already a zip archive).

All output files get chmod 600, directory gets chmod 700.

USAGE:
    python scripts/emergency_backup.py
"""

import csv
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl")
    print("    pip install openpyxl")
    sys.exit(1)


REPO_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = REPO_ROOT / "never-push-passwords" / "chrome_profile_inventory.csv"
ZIP_PATH = REPO_ROOT / "tmp" / "passwords.zip"
OUTPUT_DIR = Path.home() / "Documents" / "bwsync"


def create_excel_backup(csv_path: Path, output_path: Path) -> None:
    """Read the CSV and write a multi-tabbed Excel workbook."""
    wb = openpyxl.Workbook()

    # --- Tab 1: Profile Inventory ---
    ws_inventory = wb.active
    ws_inventory.title = "Profile Inventory"

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws_inventory.append(row)

    # Auto-width columns (approximate)
    for col in ws_inventory.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_inventory.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # --- Tab 2: Metadata ---
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Field", "Value"])
    ws_meta.append(["Backup Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_meta.append(["Source File", str(csv_path)])

    # Count data rows (minus header)
    row_count = ws_inventory.max_row - 1
    ws_meta.append(["Profile Count", row_count])
    ws_meta.append(["Script", "scripts/emergency_backup.py"])
    ws_meta.append(["Note", "Passwords.zip copied separately as-is"])

    wb.save(output_path)


def encrypt_excel(input_path: Path, password: str) -> None:
    """Encrypt an Excel file in-place using msoffcrypto-tool."""
    try:
        import msoffcrypto
    except ImportError:
        print("WARNING: msoffcrypto-tool not installed, skipping encryption.")
        print("    pip install msoffcrypto-tool")
        return

    import tempfile

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    try:
        with open(input_path, "rb") as f_in:
            file = msoffcrypto.OfficeFile(f_in)
            file.load_key(password=password)
            with open(tmp_path, "wb") as f_out:
                file.save(f_out)
        # msoffcrypto.OfficeFile expects an already-encrypted file for load_key
        # For encrypting a plain file, we need a different approach
    except Exception:
        os.unlink(tmp_path)
        # Use the correct API: encrypt a plain file
        pass

    # msoffcrypto-tool encrypts via its OfficeFile class differently:
    # We need to create an encrypted copy from scratch
    os.unlink(tmp_path)

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    try:
        with open(input_path, "rb") as f_in, open(tmp_path, "wb") as f_out:
            file = msoffcrypto.OfficeFile(f_in)
            file.load_key(password="")  # no password on source
            file.encrypt(password, f_out)
        shutil.move(tmp_path, input_path)
        print(f"    Encrypted with BWSYNC_EXCEL_PASSWORD")
    except Exception as e:
        print(f"    WARNING: Encryption failed ({e}), file saved unencrypted.")
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def main():
    print()
    print("Emergency Backup — moving sensitive data out of repo")
    print("=" * 55)
    print()

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    os.chmod(OUTPUT_DIR, 0o700)
    print(f"  Output directory: {OUTPUT_DIR}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # --- Excel backup from CSV ---
    if CSV_PATH.exists():
        excel_name = f"bwsync_emergency_backup_{timestamp}.xlsx"
        excel_path = OUTPUT_DIR / excel_name

        print(f"\n  Creating Excel backup from CSV...")
        print(f"    Source: {CSV_PATH}")
        create_excel_backup(CSV_PATH, excel_path)
        os.chmod(excel_path, 0o600)
        print(f"    Output: {excel_path}")

        # Encrypt if password is set
        password = os.environ.get("BWSYNC_EXCEL_PASSWORD", "")
        if password:
            encrypt_excel(excel_path, password)
        else:
            print("    (No BWSYNC_EXCEL_PASSWORD set — saved unencrypted)")
    else:
        print(f"  SKIP: CSV not found at {CSV_PATH}")

    # --- Copy passwords.zip ---
    if ZIP_PATH.exists():
        zip_dest = OUTPUT_DIR / f"passwords_{timestamp}.zip"
        print(f"\n  Copying passwords.zip...")
        print(f"    Source: {ZIP_PATH}")
        shutil.copy2(ZIP_PATH, zip_dest)
        os.chmod(zip_dest, 0o600)
        print(f"    Output: {zip_dest}")
    else:
        print(f"\n  SKIP: passwords.zip not found at {ZIP_PATH}")

    print()
    print("  Done. Verify the backup, then delete originals from repo:")
    print(f"    rm {CSV_PATH}")
    print(f"    rm {ZIP_PATH}")
    print()


if __name__ == "__main__":
    main()
